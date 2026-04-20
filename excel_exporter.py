"""
Module xuất file Excel tài khoản phụ huynh.
Tạo file .xlsx từ danh sách học sinh API với các cột:
STT, Lớp, Họ và tên, Tài khoản, Mật khẩu.

Hỗ trợ 2 chế độ xuất tài khoản:
- Theo Gmail: lấy email → bỏ @gmail.com → tài khoản
- Theo SĐT: lấy contactsMobile1 (fallback contactsMobile2) → tài khoản
"""

import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Các field có thể chứa email/gmail trong response API KeyTech
# Thứ tự ưu tiên: field rõ ràng nhất trước
_EMAIL_FIELD_CANDIDATES = [
    'gmail',
    'email',
    'staffEmail',
    'staffGmail',
    'mail',
    'userEmail',
    'loginName',
    'userName',
    'account',
    'staffAccount',
]

# Fields SĐT phụ huynh trong API KeyTech (ưu tiên contactsMobile1, fallback contactsMobile2)
_PHONE_FIELD_PRIMARY = 'contactsMobile1'
_PHONE_FIELD_FALLBACK = 'contactsMobile2'

DEFAULT_PASSWORD = '123456'

# Enum-like constants cho chế độ xuất
EXPORT_MODE_GMAIL = 'gmail'
EXPORT_MODE_PHONE = 'phone'


def find_email_field(students: list) -> Optional[str]:
    """
    Tự động phát hiện field chứa email trong danh sách student records.
    Duyệt qua các field ưu tiên, nếu tìm thấy field có giá trị chứa '@'
    thì trả về tên field đó.

    Returns:
        Tên field chứa email, hoặc None nếu không tìm thấy.
    """
    if not students:
        return None

    # Thử từng candidate field theo thứ tự ưu tiên
    for field in _EMAIL_FIELD_CANDIDATES:
        for student in students:
            value = student.get(field)
            if value and isinstance(value, str) and '@' in value:
                logger.info("Phát hiện email field: '%s' (ví dụ: %s)", field, value)
                return field

    # Fallback: duyệt tất cả keys trong record đầu tiên, tìm field chứa '@'
    for student in students[:min(10, len(students))]:
        for key, value in student.items():
            if isinstance(value, str) and '@' in value:
                logger.info(
                    "Phát hiện email field (fallback scan): '%s' (ví dụ: %s)",
                    key, value,
                )
                return key

    return None


def account_from_email(email: Optional[str]) -> tuple[str, Optional[str]]:
    """
    Trích tài khoản từ email.

    Rules:
    - Nếu email rỗng/None → trả về ('', warning)
    - Nếu email kết thúc bằng @gmail.com → bỏ @gmail.com
    - Nếu email có @ nhưng không phải @gmail.com → lấy phần trước @, cảnh báo
    - Nếu email không có @ → trả về nguyên chuỗi, cảnh báo

    Returns:
        (account: str, warning: Optional[str])
    """
    if not email or not isinstance(email, str) or not email.strip():
        return '', 'Thiếu email'

    email = email.strip()

    if email.lower().endswith('@gmail.com'):
        return email[:email.lower().rfind('@gmail.com')], None

    if '@' in email:
        account = email.split('@')[0]
        return account, f"Email không phải @gmail.com: {email}"

    return email, f"Email không có @: {email}"


def _get_phone_from_student(student: dict) -> tuple[str, Optional[str]]:
    """
    Lấy SĐT phụ huynh từ student record.
    Ưu tiên contactsMobile1, fallback sang contactsMobile2.

    Returns:
        (phone: str, warning: Optional[str])
    """
    # Thử contactsMobile1 trước
    phone1 = (student.get(_PHONE_FIELD_PRIMARY) or '').strip()
    if phone1:
        return phone1, None

    # Fallback: contactsMobile2
    phone2 = (student.get(_PHONE_FIELD_FALLBACK) or '').strip()
    if phone2:
        return phone2, f"Dùng {_PHONE_FIELD_FALLBACK} (không có {_PHONE_FIELD_PRIMARY})"

    return '', 'Thiếu SĐT phụ huynh'


def build_parent_account_rows(
    students: list,
    email_field: Optional[str] = None,
    on_log: Optional[callable] = None,
) -> list[dict]:
    """
    Chuyển danh sách student records từ API thành rows cho Excel.

    Args:
        students: Danh sách dict student từ API.
        email_field: Tên field chứa email. Nếu None, sẽ tự detect.
        on_log: Callback ghi log (optional).

    Returns:
        List các dict với keys: stt, class_name, full_name, account, password
    """
    def log(msg):
        logger.info(msg)
        if on_log:
            on_log(msg)

    if not students:
        log("⚠ Không có dữ liệu học sinh để xử lý.")
        return []

    # Tự detect email field nếu chưa chỉ định
    if email_field is None:
        email_field = find_email_field(students)
        if email_field:
            log(f"📧 Phát hiện trường email: '{email_field}'")
        else:
            log("⚠ Không tìm thấy trường email/gmail trong dữ liệu API.")
            log("ℹ Cột Tài khoản sẽ để trống. Hãy kiểm tra lại response API.")

    rows = []
    warning_count = 0

    for idx, student in enumerate(students, 1):
        staff_name = (student.get('staffName') or '').strip()
        dept_name = (student.get('deptName') or '').strip().strip('/')

        # Lấy email
        raw_email = (student.get(email_field) or '') if email_field else ''
        account, warning = account_from_email(raw_email)

        if warning:
            warning_count += 1
            log(f"⚠ [{idx}] {staff_name}: {warning}")

        rows.append({
            'stt': idx,
            'class_name': dept_name,
            'full_name': staff_name,
            'account': account,
            'password': DEFAULT_PASSWORD,
        })

    if warning_count > 0:
        log(f"⚠ Tổng cộng {warning_count}/{len(students)} dòng có cảnh báo về email.")

    return rows


def build_parent_account_rows_by_phone(
    students: list,
    on_log: Optional[callable] = None,
) -> list[dict]:
    """
    Chuyển danh sách student records từ API thành rows cho Excel.
    Tài khoản lấy từ SĐT phụ huynh (contactsMobile1, fallback contactsMobile2).

    Args:
        students: Danh sách dict student từ API.
        on_log: Callback ghi log (optional).

    Returns:
        List các dict với keys: stt, class_name, full_name, account, password
    """
    def log(msg):
        logger.info(msg)
        if on_log:
            on_log(msg)

    if not students:
        log("⚠ Không có dữ liệu học sinh để xử lý.")
        return []

    # Kiểm tra xem API có trả về các field SĐT không
    first = students[0]
    has_primary = _PHONE_FIELD_PRIMARY in first
    has_fallback = _PHONE_FIELD_FALLBACK in first

    if has_primary:
        log(f"📱 Tìm thấy trường SĐT: '{_PHONE_FIELD_PRIMARY}'")
    if has_fallback:
        log(f"📱 Tìm thấy trường SĐT dự phòng: '{_PHONE_FIELD_FALLBACK}'")
    if not has_primary and not has_fallback:
        log(f"⚠ Không tìm thấy trường '{_PHONE_FIELD_PRIMARY}' hoặc '{_PHONE_FIELD_FALLBACK}' trong API.")
        log("ℹ Cột Tài khoản sẽ để trống.")

    rows = []
    warning_count = 0
    fallback_count = 0
    missing_count = 0

    for idx, student in enumerate(students, 1):
        staff_name = (student.get('staffName') or '').strip()
        dept_name = (student.get('deptName') or '').strip().strip('/')

        # Lấy SĐT: ưu tiên contactsMobile1, fallback contactsMobile2
        account, warning = _get_phone_from_student(student)

        if warning:
            if 'Thiếu' in warning:
                missing_count += 1
            else:
                fallback_count += 1
            warning_count += 1
            log(f"⚠ [{idx}] {staff_name}: {warning}")

        rows.append({
            'stt': idx,
            'class_name': dept_name,
            'full_name': staff_name,
            'account': account,
            'password': DEFAULT_PASSWORD,
        })

    # Tổng kết
    ok_count = len(students) - missing_count
    log(f"📊 Kết quả: {ok_count}/{len(students)} có SĐT"
        f"{f' ({fallback_count} dùng {_PHONE_FIELD_FALLBACK})' if fallback_count > 0 else ''}"
        f"{f', {missing_count} thiếu SĐT' if missing_count > 0 else ''}")

    return rows


def export_parent_accounts_to_excel(
    rows: list[dict],
    output_path: str,
    on_log: Optional[callable] = None,
) -> bool:
    """
    Xuất danh sách tài khoản phụ huynh ra file Excel (.xlsx).

    Args:
        rows: List dict từ build_parent_account_rows().
        output_path: Đường dẫn file .xlsx output.
        on_log: Callback ghi log (optional).

    Returns:
        True nếu thành công, False nếu lỗi.
    """
    def log(msg):
        logger.info(msg)
        if on_log:
            on_log(msg)

    if not rows:
        log("⚠ Không có dữ liệu để xuất Excel.")
        return False

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Tài khoản phụ huynh"

        # === Header ===
        headers = ['STT', 'Lớp', 'Họ và tên', 'Tài khoản', 'Mật khẩu']
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2D8A4E', end_color='2D8A4E', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC'),
        )

        for col_idx, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # === Data rows ===
        data_font = Font(name='Arial', size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')

        for row_data in rows:
            row_num = row_data['stt'] + 1  # +1 vì row 1 là header
            values = [
                row_data['stt'],
                row_data['class_name'],
                row_data['full_name'],
                row_data['account'],
                row_data['password'],
            ]
            alignments = [center_align, center_align, left_align, left_align, center_align]

            for col_idx, (value, align) in enumerate(zip(values, alignments), 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = align
                cell.border = thin_border

        # === Auto-fit column widths ===
        column_widths = {
            1: 6,    # STT
            2: 15,   # Lớp
            3: 30,   # Họ và tên
            4: 25,   # Tài khoản
            5: 12,   # Mật khẩu
        }

        # Tính width thực tế dựa trên dữ liệu
        for row_data in rows:
            values = [
                str(row_data['stt']),
                row_data['class_name'],
                row_data['full_name'],
                row_data['account'],
                row_data['password'],
            ]
            for col_idx, val in enumerate(values, 1):
                needed = len(str(val)) + 3
                if needed > column_widths.get(col_idx, 0):
                    column_widths[col_idx] = min(needed, 50)  # cap at 50

        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # === Freeze header row ===
        ws.freeze_panes = 'A2'

        # === Save ===
        wb.save(output_path)
        log(f"✅ Đã xuất file Excel: {output_path}")
        log(f"📊 Tổng cộng {len(rows)} dòng dữ liệu.")
        return True

    except PermissionError:
        log(f"❌ Không thể lưu file — file đang được mở bởi chương trình khác: {output_path}")
        return False
    except Exception as e:
        log(f"❌ Lỗi khi xuất Excel: {e}")
        logger.exception("Excel export error")
        return False
